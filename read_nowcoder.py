import openpyxl
import sys

sys.stdout.reconfigure(encoding='utf-8')

wb = openpyxl.load_workbook('data/nowcoder_contest1.xlsx')
ws = wb.active
print(f'Sheet name: {ws.title}')
print(f'Max row: {ws.max_row}, Max col: {ws.max_column}')
print('\nHeaders:')
for cell in ws[1]:
    print(f'  {cell.column_letter}: {cell.value}')

print('\nAll rows:')
for row in ws.iter_rows(min_row=2, values_only=True):
    print(row)