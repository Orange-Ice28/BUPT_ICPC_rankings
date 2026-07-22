import json
import os
import re
import openpyxl

# Resolve project root (script lives in scripts/)
os.chdir(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

DATA_DIR = "data"
SCORE_FILE = f"{DATA_DIR}/score_result.xlsx"
RANK_FILE = f"{DATA_DIR}/hdu_rank_all.xlsx"
BASELINE_FILE = f"{DATA_DIR}/baseline.xlsx"
TEAM_FILE = f"{DATA_DIR}/team.xlsx"

NUM_CONTESTS = 10
BEST_N_OLD = 7
BEST_N_NEW = 5
TEAM_THRESHOLD = 1791


def load_baseline():
    wb = openpyxl.load_workbook(BASELINE_FILE)
    ws = wb.active
    baselines = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        baselines.append(int(row[1]))
    wb.close()
    return baselines


def load_teams():
    wb = openpyxl.load_workbook(TEAM_FILE)
    ws = wb.active
    teams = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        seq, name_cn, name_en, leader, m1, m2 = row
        if name_cn is None:
            continue
        members = [m for m in (leader, m1, m2) if m is not None]
        teams.append({
            "seq": seq,
            "name_cn": name_cn,
            "name_en": name_en,
            "members": members,
        })
    wb.close()
    return teams


def load_rank_data():
    wb = openpyxl.load_workbook(RANK_FILE)
    ws = wb.active
    data = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        team_id = row[0]
        name = row[1]
        contests = []
        for i in range(NUM_CONTESTS):
            solved = row[2 + i * 2]
            rank = row[3 + i * 2]
            if solved is None:
                solved = 0
            if rank is None:
                rank = 0
            contests.append({
                "solved": int(solved),
                "rank": int(rank),
            })
        data.append({
            "team_id": team_id,
            "name": name,
            "contests": contests,
        })
    wb.close()
    return data


def calc_contest_score(solved, rank, baseline):
    if baseline == 0:
        return 0.0
    score = (solved / baseline) * (801 - rank) / 800 * 100
    if score < 0 or score > 100:
        return 0.0
    return score


def calc_personal_total(scores, team_id):
    match = re.search(r"team(\d+)", team_id)
    num = int(match.group(1)) if match else 0
    best_n = BEST_N_OLD if num <= TEAM_THRESHOLD else BEST_N_NEW
    pairs = [(scores[i], i) for i in range(len(scores))]
    if not pairs:
        return 0.0, set()
    pairs.sort(key=lambda x: x[0], reverse=True)
    top = pairs[:best_n]
    best_indices = set(p[1] for p in top)
    total = sum(p[0] for p in top) / best_n
    return total, best_indices


def main():
    baselines = load_baseline()
    teams = load_teams()
    rank_data = load_rank_data()
    rank_data = [item for item in rank_data if item["team_id"] != "team1790"]

    name_to_team_id = {}
    for item in rank_data:
        name_to_team_id[item["name"]] = item["team_id"]

    team_id_to_data = {}
    for item in rank_data:
        team_id_to_data[item["team_id"]] = item

    team_member_names = set()
    for team in teams:
        for m in team["members"]:
            team_member_names.add(m)

    personal_results = []
    for item in rank_data:
        tid = item["team_id"]
        scores = []
        contest_details = []
        for i, c in enumerate(item["contests"]):
            s = calc_contest_score(c["solved"], c["rank"], baselines[i])
            scores.append(s)
            contest_details.append({
                "solved": c["solved"],
                "rank": c["rank"],
                "score": round(s, 2),
            })

        total, best_indices = calc_personal_total(scores, tid)
        personal_results.append({
            "team_id": tid,
            "name": item["name"],
            "total_score": round(total, 2),
            "contests": contest_details,
            "best_indices": sorted(list(best_indices)),
            "in_team": item["name"] in team_member_names,
        })

    zhou_team_id = name_to_team_id.get("周弋然")
    if zhou_team_id:
        for p in personal_results:
            if p["team_id"] == zhou_team_id:
                p["contests"][7]["score"] = 0.0
                scores = [d["score"] for d in p["contests"]]
                total, best_indices = calc_personal_total(scores, zhou_team_id)
                p["total_score"] = round(total, 2)
                p["best_indices"] = sorted(list(best_indices))
                break

    personal_results.sort(key=lambda x: x["total_score"], reverse=True)
    for i, p in enumerate(personal_results):
        p["rank"] = i + 1

    team_results = []
    for team in teams:
        member_info = []
        member_scores = []
        for member_name in team["members"]:
            tid = name_to_team_id.get(member_name)
            if tid:
                pr = next((p for p in personal_results if p["team_id"] == tid), None)
                if pr:
                    member_info.append({
                        "name": member_name,
                        "total_score": pr["total_score"],
                    })
                    member_scores.append(pr["total_score"])
                else:
                    member_info.append({"name": member_name, "total_score": 0.0})
                    member_scores.append(0.0)
            else:
                member_info.append({"name": member_name, "total_score": 0.0})
                member_scores.append(0.0)

        team_total = sum(member_scores) / len(member_scores) if member_scores else 0.0
        team_results.append({
            "name_cn": team["name_cn"],
            "name_en": team["name_en"],
            "members": member_info,
            "team_total": round(team_total, 2),
        })

    team_results.sort(key=lambda x: x["team_total"], reverse=True)
    for i, t in enumerate(team_results):
        t["rank"] = i + 1

    result = {
        "baselines": baselines,
        "personal": personal_results,
        "teams": team_results,
    }

    with open(f"{DATA_DIR}/score_data.json", "w", encoding="utf-8") as f:
        json.dump(result, f, ensure_ascii=False, indent=2)

    with open(f"{DATA_DIR}/export_log.txt", "w", encoding="utf-8") as f:
        f.write(f"Personal count: {len(personal_results)}\n")
        f.write(f"Team count: {len(team_results)}\n")
        f.write(f"Baselines: {baselines}\n")
        f.write("Done.\n")


if __name__ == "__main__":
    main()