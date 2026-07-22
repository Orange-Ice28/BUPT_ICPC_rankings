"""
暑期集训成绩更新脚本（统一版：牛客 + 杭电）
=============================================
用法:
  python scripts/update_summer.py nc 3       # 更新牛客第3场
  python scripts/update_summer.py nc 3 10    # 更新牛客第3场，指定 baseline=10
  python scripts/update_summer.py hdu 1      # 更新杭电第1场
  python scripts/update_summer.py hdu 2 7    # 更新杭电第2场，指定 baseline=7

功能:
  1. 读取对应 Excel sheet 的队伍数据
  2. 按对应公式计算每队得分，填回 Excel
  3. 更新 data/summer_score_data.json
  4. 重新计算队伍总成绩（best 80% 规则）和排名
  5. 同步到 bupt-ranking/ 前端目录
"""

import openpyxl
import json
import math
import os
import sys
import shutil

sys.stdout.reconfigure(encoding='utf-8')

# Resolve project root (script lives in scripts/)
os.chdir(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

# ============================================================
# 比赛类型配置
# ============================================================
CONTEST_CONFIG = {
    "nc": {
        "label": "牛客",
        "excel_file": "data/nowcoder_contest.xlsx",
        "sheet_template": "nc{num}",       # nc1, nc2, ...
        "index_base": 0,                    # nc1 → summer[0], nc2 → summer[1], ...
        "rank_offset": 601,                 # 公式中的排名上限参数
        "default_baseline": 9,
        # Excel 列映射（0-based）: 姓名, 学校, 过题数, 排名, Baseline, 得分
        "col_name": 0,
        "col_solved": 2,
        "col_rank": 3,
        "col_baseline": 4,
        "col_score": 5,
    },
    "hdu": {
        "label": "杭电",
        "excel_file": "data/hdu_contest.xlsx",
        "sheet_template": "hdu{num}",       # hdu1, hdu2, ...
        "index_base": 10,                   # hdu1 → summer[10], hdu2 → summer[11], ...
        "rank_offset": 501,                 # 公式中的排名上限参数
        "default_baseline": 7,
        # Excel 列映射（0-based）: 姓名, 学校, 排名, 过题数, Baseline, 得分
        "col_name": 0,
        "col_solved": 3,                    # HDU: 过题数在第4列
        "col_rank": 2,                      # HDU: 排名在第3列
        "col_baseline": 4,
        "col_score": 5,
    },
}

DATA_DIR = "data"
SUMMER_DATA_FILE = os.path.join(DATA_DIR, "summer_score_data.json")
FRONTEND_DIRS = [
    os.path.join("bupt-ranking", "src", "data"),
    os.path.join("bupt-ranking", "public"),
]

# 队伍名映射（Excel 中的名字 → summer_score_data.json 中的 name_cn）
NAME_MAPPING = {
    # 示例: 'Excel中的名字': 'JSON中的名字',
}

# ============================================================
# 工具函数
# ============================================================

def calc_score(solved, rank, baseline, rank_offset):
    """计算暑期训练单场得分
    通用公式: 得分 = 过题数 / baseline × (rank_offset − 排名) / (rank_offset − 1) × 100
    得分 clamp 到 [0, 100]
    """
    if baseline == 0 or rank_offset <= 1:
        return 0.0
    score = (solved / baseline) * (rank_offset - rank) / (rank_offset - 1) * 100
    if score < 0 or score > 100:
        return 0.0
    return round(score, 2)


def load_excel_teams(config, sheet_name):
    """从 Excel 读取队伍数据，根据配置映射列"""
    excel_file = config["excel_file"]

    if not os.path.exists(excel_file):
        print(f"错误: 找不到文件 '{excel_file}'")
        return None, None, None

    wb = openpyxl.load_workbook(excel_file)

    if sheet_name not in wb.sheetnames:
        print(f"错误: 找不到 sheet '{sheet_name}'")
        print(f"可用的 sheet: {wb.sheetnames}")
        wb.close()
        return None, None, None

    ws = wb[sheet_name]
    teams = []

    cn = config["col_name"]
    cs = config["col_solved"]
    cr = config["col_rank"]
    cb = config["col_baseline"]

    for row in ws.iter_rows(min_row=2, values_only=True):
        name = row[cn] if cn < len(row) else None
        if name is None:
            continue
        teams.append({
            "name": name,
            "solved": int(row[cs]) if cs < len(row) and row[cs] else 0,
            "rank": int(row[cr]) if cr < len(row) and row[cr] else 0,
            "baseline": int(row[cb]) if cb < len(row) and row[cb] else 0,
            "score": row[config["col_score"]] if config["col_score"] < len(row) else None,
        })

    return wb, ws, teams


def load_summer_data():
    """加载 summer_score_data.json"""
    with open(SUMMER_DATA_FILE, "r", encoding="utf-8") as f:
        return json.load(f)


def save_summer_data(data):
    """保存 summer_score_data.json 并同步到前端"""
    with open(SUMMER_DATA_FILE, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    print(f"  已保存: {SUMMER_DATA_FILE}")

    for d in FRONTEND_DIRS:
        target = os.path.join(d, "summer_score_data.json")
        os.makedirs(d, exist_ok=True)
        shutil.copy(SUMMER_DATA_FILE, target)
        print(f"  已同步: {target}")


def match_team(excel_name, summer_teams):
    """将 Excel 中的队伍名匹配到 summer_score_data.json 中的队伍"""
    # 先检查映射表
    if excel_name in NAME_MAPPING:
        mapped = NAME_MAPPING[excel_name]
        for t in summer_teams:
            if t["name_cn"] == mapped:
                return t

    # 精确匹配中文名
    for t in summer_teams:
        if t["name_cn"] == excel_name:
            return t

    # 模糊匹配中文名
    for t in summer_teams:
        if excel_name in t["name_cn"] or t["name_cn"] in excel_name:
            return t

    # 匹配英文名（标准化后比较）
    def normalize(s):
        return s.replace(" ", "").replace("://", "").replace("!", "").lower()

    nc_norm = normalize(excel_name)
    for t in summer_teams:
        en_norm = normalize(t.get("name_en", ""))
        if en_norm and (nc_norm == en_norm or en_norm in nc_norm or nc_norm in en_norm):
            return t

    return None


def recalc_team_totals(data):
    """重新计算所有队伍的总成绩和排名（best 80% 规则）"""
    for team in data["teams"]:
        contests_with_data = sum(
            1 for c in team["contests"] if c["solved"] > 0 or c["score"] > 0
        )

        # 重置所有 isBest
        for contest in team["contests"]:
            contest["isBest"] = False

        if contests_with_data == 0:
            team["team_total"] = 0.0
        elif contests_with_data == 1:
            for i, c in enumerate(team["contests"]):
                if c["solved"] > 0 or c["score"] > 0:
                    team["team_total"] = round(c["score"], 2)
                    c["isBest"] = True
                    break
        else:
            best_n = math.ceil(contests_with_data * 0.8)
            pairs = [
                (i, c["score"])
                for i, c in enumerate(team["contests"])
                if c["solved"] > 0 or c["score"] > 0
            ]
            pairs.sort(key=lambda x: x[1], reverse=True)
            best_pairs = pairs[:best_n]
            best_indices = set(p[0] for p in best_pairs)

            for i in best_indices:
                team["contests"][i]["isBest"] = True

            best_scores = [p[1] for p in best_pairs]
            team["team_total"] = round(sum(best_scores) / best_n, 2)

    # 排序并更新排名
    data["teams"].sort(key=lambda x: x["team_total"], reverse=True)
    for i, team in enumerate(data["teams"]):
        team["rank"] = i + 1


# ============================================================
# 主流程
# ============================================================

def main():
    # 解析命令行参数
    if len(sys.argv) < 2:
        print("用法: python scripts/update_summer.py <类型> <场次> [baseline]")
        print()
        print("类型:")
        print("  nc    牛客比赛 (sheet: nc{num}, 索引: 0~9)")
        print("  hdu   杭电比赛 (sheet: hdu{num}, 索引: 10~19)")
        print()
        print("示例:")
        print("  python scripts/update_summer.py nc 3       # 牛客第3场")
        print("  python scripts/update_summer.py nc 3 10    # 牛客第3场，baseline=10")
        print("  python scripts/update_summer.py hdu 1      # 杭电第1场")
        print("  python scripts/update_summer.py hdu 2 7    # 杭电第2场，baseline=7")
        sys.exit(1)

    contest_type = sys.argv[1].lower()

    if contest_type not in CONTEST_CONFIG:
        print(f"错误: 未知比赛类型 '{contest_type}'，支持: {list(CONTEST_CONFIG.keys())}")
        sys.exit(1)

    try:
        contest_num = int(sys.argv[2])
    except (ValueError, IndexError):
        print("错误: 场次必须是整数")
        sys.exit(1)

    config = CONTEST_CONFIG[contest_type]
    sheet_name = config["sheet_template"].format(num=contest_num)
    contest_idx = config["index_base"] + contest_num - 1
    default_baseline = int(sys.argv[3]) if len(sys.argv) >= 4 else None
    rank_offset = config["rank_offset"]

    print("=" * 65)
    print(f"  更新{config['label']}第 {contest_num} 场 (sheet: {sheet_name}, idx: {contest_idx})")
    print("=" * 65)
    print()

    # ---- 第1步：读取 Excel 数据 ----
    print("[1/5] 读取 Excel 数据...")
    wb, ws, excel_teams = load_excel_teams(config, sheet_name)
    if excel_teams is None:
        sys.exit(1)
    print(f"  从 {config['excel_file']} / {sheet_name} 读取到 {len(excel_teams)} 支队伍")

    # ---- 第2步：计算得分 ----
    print()
    formula_desc = f"(solved/baseline) x ({rank_offset} - rank) / {rank_offset - 1} x 100"
    print(f"[2/5] 计算得分 — 公式: {formula_desc}")
    print()

    for team in excel_teams:
        baseline = default_baseline or team["baseline"] or config["default_baseline"]
        if team["baseline"] == 0 or team["baseline"] is None:
            team["baseline"] = baseline

        score = calc_score(team["solved"], team["rank"], baseline, rank_offset)
        team["score"] = score
        team["baseline"] = baseline
        print(f"  {team['name']:<35s}  "
              f"solved={team['solved']:<3d}  rank={team['rank']:<6d}  "
              f"baseline={baseline}  →  score={score:>7.2f}")

    # ---- 第3步：写回 Excel ----
    print()
    print("[3/5] 写回 Excel...")
    score_col = config["col_score"] + 1  # openpyxl 1-indexed
    baseline_col = config["col_baseline"] + 1

    for i, team in enumerate(excel_teams):
        row_idx = i + 2  # 第1行是表头
        ws.cell(row=row_idx, column=score_col).value = team["score"]
        existing_baseline = ws.cell(row=row_idx, column=baseline_col).value
        if existing_baseline is None or existing_baseline == 0:
            ws.cell(row=row_idx, column=baseline_col).value = team["baseline"]

    wb.save(config["excel_file"])
    wb.close()
    print(f"  已保存到: {config['excel_file']}")

    # ---- 第4步：更新 summer_score_data.json ----
    print()
    print("[4/5] 更新 summer_score_data.json...")
    data = load_summer_data()

    # 确保 baselines 数组足够长
    while len(data["baselines"]) <= contest_idx:
        data["baselines"].append(0)

    used_baseline = excel_teams[0]["baseline"] if excel_teams else config["default_baseline"]
    data["baselines"][contest_idx] = used_baseline
    print(f"  baselines[{contest_idx}] = {used_baseline}")

    # 更新每支队伍的 contest 数据
    matched_count = 0
    unmatched = []

    for excel_team in excel_teams:
        matched = match_team(excel_team["name"], data["teams"])
        if matched:
            matched["contests"][contest_idx] = {
                "solved": excel_team["solved"],
                "rank": excel_team["rank"],
                "score": excel_team["score"],
                "isBest": False,
            }
            matched_count += 1
            print(f"  OK '{excel_team['name']}' → '{matched['name_cn']}' "
                  f"(solved={excel_team['solved']}, rank={excel_team['rank']}, score={excel_team['score']})")
        else:
            unmatched.append(excel_team["name"])
            print(f"  !! '{excel_team['name']}' → 未找到匹配队伍!")

    if unmatched:
        print()
        print(f"  警告: {len(unmatched)} 支队伍未匹配:")
        for name in unmatched:
            print(f"    - {name}")
        print(f"  请在脚本顶部的 NAME_MAPPING 中添加映射。")

    # ---- 第5步：重新计算总成绩并保存 ----
    print()
    print("[5/5] 重新计算队伍总成绩...")
    recalc_team_totals(data)

    # 打印排名
    print()
    print("  Top 10 队伍总成绩:")
    for team in data["teams"][:10]:
        n_data = sum(1 for c in team["contests"] if c["solved"] > 0 or c["score"] > 0)
        print(f"    {team['rank']:>2}. {team['name_cn']:<20s}  "
              f"{team['team_total']:>6.2f}  (有效场次: {n_data})")

    save_summer_data(data)

    # ---- 完成 ----
    print()
    print("=" * 65)
    print(f"  {config['label']}第 {contest_num} 场更新完成! 共处理 {len(excel_teams)} 支队伍")
    print("=" * 65)


if __name__ == "__main__":
    main()
