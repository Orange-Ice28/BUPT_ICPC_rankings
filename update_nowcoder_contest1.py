import openpyxl
import sys
import json
import os

sys.stdout.reconfigure(encoding='utf-8')

DATA_DIR = "data"
SCORE_DATA_FILE = os.path.join(DATA_DIR, "score_data.json")
NOWCODER_FILE = os.path.join(DATA_DIR, "nowcoder_contest1.xlsx")
TEAM_FILE = os.path.join(DATA_DIR, "team.xlsx")

BASELINE = 9  # 牛客第一场 baseline 为 9 题

def load_score_data():
    """加载 score_data.json"""
    with open(SCORE_DATA_FILE, "r", encoding="utf-8") as f:
        return json.load(f)

def load_teams():
    """加载队伍信息"""
    wb = openpyxl.load_workbook(TEAM_FILE)
    ws = wb.active
    teams = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        seq, name_cn, name_en, leader, m1, m2 = row
        if name_cn is None:
            continue
        members = [m for m in (leader, m1, m2) if m is not None]
        teams.append({
            "seq": seq,
            "name_cn": name_cn,
            "name_en": name_en,
            "members": members,
        })
    wb.close()
    return teams

def calc_summer_score(solved, rank, baseline):
    """计算暑期训练单场得分
    公式: 得分 = 过题数 / baseline × (601 − 排名) / 600 × 100
    """
    if baseline == 0:
        return 0.0
    score = (solved / baseline) * (601 - rank) / 600 * 100
    if score < 0 or score > 100:
        return 0.0
    return round(score, 2)

def update_nowcoder_excel():
    """更新 nowcoder_contest1.xlsx，添加 nullPointerException 队伍的成绩"""
    wb = openpyxl.load_workbook(NOWCODER_FILE)
    ws = wb.active
    
    print(f"当前表格: {ws.title}")
    print(f"当前行数: {ws.max_row}, 列数: {ws.max_column}")
    
    # 添加 nullPointerException 队伍的成绩
    # 5题244名，baseline=9
    new_row = {
        "rank": "244",
        "name": "nullPointerException",
        "school": "北京邮电大学",
        "solved": "5",
        "penalty": "",  # 罚时未知，留空
        "baseline": BASELINE,
        "score": calc_summer_score(5, 244, BASELINE)
    }
    
    print(f"\n添加新记录:")
    print(f"  排名: {new_row['rank']}")
    print(f"  姓名: {new_row['name']}")
    print(f"  学校: {new_row['school']}")
    print(f"  过题数: {new_row['solved']}")
    print(f"  得分: {new_row['score']}")
    
    # 添加到表格末尾
    ws.append([
        new_row["rank"],
        new_row["name"],
        new_row["school"],
        new_row["solved"],
        new_row["penalty"],
        new_row["baseline"],
        new_row["score"]
    ])
    
    # 保存
    wb.save(NOWCODER_FILE)
    print(f"\n已保存到: {NOWCODER_FILE}")
    print(f"新表格行数: {ws.max_row}")

def main():
    print("=" * 60)
    print("更新 nowcoder_contest1.xlsx")
    print("=" * 60)
    
    # 更新 Excel 表格
    update_nowcoder_excel()
    
    print("\n" + "=" * 60)
    print("更新完成!")
    print("=" * 60)

if __name__ == "__main__":
    main()