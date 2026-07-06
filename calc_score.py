import os
import re
from collections import defaultdict

import openpyxl
from openpyxl.styles import PatternFill

DATA_DIR = "data"
RANK_FILE = os.path.join(DATA_DIR, "hdu_rank_all.xlsx")
BASELINE_FILE = os.path.join(DATA_DIR, "baseline.xlsx")
TEAM_FILE = os.path.join(DATA_DIR, "team.xlsx")
OUTPUT_FILE = os.path.join(DATA_DIR, "score_result.xlsx")

NUM_CONTESTS = 10
BEST_N_OLD = 7
BEST_N_NEW = 5
TEAM_THRESHOLD = 1791

GRAY_FILL = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")


def load_rank_data():
    wb = openpyxl.load_workbook(RANK_FILE)
    ws = wb.active
    data = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        team_id = row[0]
        name = row[1]
        contests = []
        for i in range(NUM_CONTESTS):
            solved = row[2 + i * 2]
            rank = row[3 + i * 2]
            if solved is None:
                solved = 0
            if rank is None:
                rank = 0
            contests.append({
                "solved": int(solved),
                "rank": int(rank),
            })
        data.append({
            "team_id": team_id,
            "name": name,
            "contests": contests,
        })
    wb.close()
    return data


def load_baseline():
    wb = openpyxl.load_workbook(BASELINE_FILE)
    ws = wb.active
    baselines = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        baselines.append(int(row[1]))
    wb.close()
    return baselines


def load_teams():
    wb = openpyxl.load_workbook(TEAM_FILE)
    ws = wb.active
    teams = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        seq, name_cn, name_en, leader, m1, m2 = row
        if name_cn is None:
            continue
        members = [m for m in (leader, m1, m2) if m is not None]
        teams.append({
            "seq": seq,
            "name_cn": name_cn,
            "name_en": name_en,
            "members": members,
        })
    wb.close()
    return teams


def calc_contest_score(solved, rank, baseline):
    if baseline == 0:
        return 0.0
    score = (solved / baseline) * (801 - rank) / 800 * 100
    if score < 0 or score > 100:
        return 0.0
    return score


def calc_personal_total(scores, team_id):
    match = re.search(r"team(\d+)", team_id)
    num = int(match.group(1)) if match else 0
    best_n = BEST_N_OLD if num <= TEAM_THRESHOLD else BEST_N_NEW

    pairs = [(scores[i], i) for i in range(len(scores)) if scores[i] > 0]
    if not pairs:
        return 0.0, set()

    pairs.sort(key=lambda x: x[0], reverse=True)
    top = pairs[:best_n]
    best_indices = set(p[1] for p in top)
    total = sum(p[0] for p in top) / len(top)
    return total, best_indices


def main():
    rank_data = load_rank_data()
    rank_data = [item for item in rank_data if item["team_id"] != "team1790"]
    baselines = load_baseline()
    teams = load_teams()

    name_to_team_id = {}
    for item in rank_data:
        name_to_team_id[item["name"]] = item["team_id"]

    team_id_to_data = {}
    for item in rank_data:
        team_id_to_data[item["team_id"]] = item

    personal_scores = {}
    personal_contest_details = {}
    personal_best_indices = {}

    for item in rank_data:
        tid = item["team_id"]
        scores = []
        details = []
        for i, c in enumerate(item["contests"]):
            s = calc_contest_score(c["solved"], c["rank"], baselines[i])
            scores.append(s)
            details.append({
                "solved": c["solved"],
                "rank": c["rank"],
                "score": round(s, 2),
            })
        total, best_indices = calc_personal_total(scores, tid)
        personal_scores[tid] = round(total, 2)
        personal_contest_details[tid] = details
        personal_best_indices[tid] = best_indices

    zhou_team_id = name_to_team_id.get("周弋然")
    if zhou_team_id:
        personal_contest_details[zhou_team_id][7]["score"] = 0.0
        scores = [d["score"] for d in personal_contest_details[zhou_team_id]]
        total, best_indices = calc_personal_total(scores, zhou_team_id)
        personal_scores[zhou_team_id] = round(total, 2)
        personal_best_indices[zhou_team_id] = best_indices

    wb = openpyxl.Workbook()

    team_member_names = set()
    for team in teams:
        for m in team["members"]:
            team_member_names.add(m)

    ws1 = wb.active
    ws1.title = "个人每场成绩"

    headers1 = ["排名", "队伍编号", "姓名", "个人总成绩"]
    for i in range(1, NUM_CONTESTS + 1):
        headers1.append(f"第{i}场 过题数")
        headers1.append(f"第{i}场 排名")
        headers1.append(f"第{i}场 得分")
    ws1.append(headers1)

    sorted_persons = sorted(team_id_to_data.keys(), key=lambda x: personal_scores[x], reverse=True)

    in_team = [tid for tid in sorted_persons if team_id_to_data[tid]["name"] in team_member_names]
    not_in_team = [tid for tid in sorted_persons if team_id_to_data[tid]["name"] not in team_member_names]

    def write_person_rows(ws, tid_list, start_rank):
        rank = start_rank
        for tid in tid_list:
            item = team_id_to_data[tid]
            row_data = [rank, tid, item["name"], personal_scores[tid]]
            rank += 1
            for d in personal_contest_details[tid]:
                row_data.append(d["solved"])
                row_data.append(d["rank"])
                row_data.append(d["score"])
            ws.append(row_data)
            row_idx = ws.max_row
            best = personal_best_indices[tid]
            for ci in range(NUM_CONTESTS):
                if ci not in best:
                    col_start = 5 + ci * 3
                    for dc in range(3):
                        ws.cell(row=row_idx, column=col_start + dc).fill = GRAY_FILL
        return rank

    rank_counter = 1
    rank_counter = write_person_rows(ws1, in_team, rank_counter)

    if not_in_team:
        ws1.append([])
        rank_counter = write_person_rows(ws1, not_in_team, rank_counter)

    col_widths_1 = {"A": 8, "B": 14, "C": 16, "D": 14}
    for col, width in col_widths_1.items():
        ws1.column_dimensions[col].width = width

    ws2 = wb.create_sheet("队伍总体情况")

    max_members = max(len(t["members"]) for t in teams)
    headers2 = ["排名", "中文队名", "英文队名"]
    for i in range(1, max_members + 1):
        headers2.append(f"队员{i}")
        headers2.append(f"队员{i}总成绩")
    headers2.append("队伍总成绩")
    ws2.append(headers2)

    team_results = []
    for team in teams:
        member_scores = []
        member_names = []
        for member_name in team["members"]:
            tid = name_to_team_id.get(member_name)
            if tid:
                member_names.append(member_name)
                member_scores.append(personal_scores.get(tid, 0.0))
            else:
                member_names.append(member_name)
                member_scores.append(0.0)

        team_total = sum(member_scores) / len(member_scores) if member_scores else 0.0
        team_results.append({
            "name_cn": team["name_cn"],
            "name_en": team["name_en"],
            "member_names": member_names,
            "member_scores": member_scores,
            "team_total": round(team_total, 2),
        })

    team_results.sort(key=lambda x: x["team_total"], reverse=True)

    for rank_idx, tr in enumerate(team_results, 1):
        row_data = [rank_idx, tr["name_cn"], tr["name_en"]]
        for i in range(max_members):
            if i < len(tr["member_names"]):
                row_data.append(tr["member_names"][i])
                row_data.append(tr["member_scores"][i])
            else:
                row_data.append("")
                row_data.append("")
        row_data.append(tr["team_total"])
        ws2.append(row_data)

    col_widths_2 = {"A": 8, "B": 28, "C": 30}
    for col, width in col_widths_2.items():
        ws2.column_dimensions[col].width = width

    os.makedirs(DATA_DIR, exist_ok=True)
    wb.save(OUTPUT_FILE)
    print(f"结果已保存到 {OUTPUT_FILE}")

    print("\n=== 队伍总成绩排名 ===")
    for rank_idx, tr in enumerate(team_results, 1):
        print(f"  {rank_idx}. {tr['name_cn']} ({tr['name_en']}): {tr['team_total']:.2f}")


if __name__ == "__main__":
    main()