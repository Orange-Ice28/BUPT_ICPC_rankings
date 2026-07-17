import openpyxl
import json
import os
import re

DATA_DIR = "data"
SCORE_DATA_FILE = os.path.join(DATA_DIR, "score_data.json")
NOWCODER_FILE = os.path.join(DATA_DIR, "nowcoder_contest1.xlsx")
TEAM_FILE = os.path.join(DATA_DIR, "team.xlsx")

BASELINE = 9
NUM_CONTESTS = 10
BEST_N_OLD = 7
BEST_N_NEW = 5
TEAM_THRESHOLD = 1791

def calc_summer_score(solved, rank, baseline):
    """计算暑期训练单场得分"""
    if baseline == 0:
        return 0.0
    score = (solved / baseline) * (601 - rank) / 600 * 100
    if score < 0 or score > 100:
        return 0.0
    return round(score, 2)

def load_score_data():
    with open(SCORE_DATA_FILE, "r", encoding="utf-8") as f:
        return json.load(f)

def load_teams():
    wb = openpyxl.load_workbook(TEAM_FILE)
    ws = wb.active
    teams = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        seq, name_cn, name_en, leader, m1, m2 = row
        if name_cn is None:
            continue
        members = [m for m in (leader, m1, m2) if m is not None]
        teams.append({
            "seq": seq,
            "name_cn": name_cn,
            "name_en": name_en,
            "members": members,
        })
    wb.close()
    return teams

def load_nowcoder_data():
    wb = openpyxl.load_workbook(NOWCODER_FILE)
    ws = wb.active
    data = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        rank, name, school, solved, penalty, baseline, score = row
        data.append({
            "rank": int(rank) if rank else 0,
            "name": name,
            "school": school,
            "solved": int(solved) if solved else 0,
            "penalty": penalty,
            "baseline": baseline,
            "score": score
        })
    wb.close()
    return data

def main():
    score_data = load_score_data()
    nowcoder_data = load_nowcoder_data()
    teams = load_teams()
    
    # 构建姓名到队伍信息的映射
    name_to_team = {}
    for team in teams:
        for member in team["members"]:
            name_to_team[member] = team
    
    # 更新每个人的牛客第一场成绩
    for person in score_data["personal"]:
        name = person["name"]
        
        # 查找该人在牛客第一场的成绩
        nowcoder_record = None
        for record in nowcoder_data:
            if record["name"] == name or name in record["name"] or record["name"] in name:
                nowcoder_record = record
                break
        
        # 如果找到记录，更新 contests[0]（牛客第一场）
        if nowcoder_record:
            person["contests"][0] = {
                "solved": nowcoder_record["solved"],
                "rank": nowcoder_record["rank"],
                "score": nowcoder_record["score"]
            }
            print(f"更新 {name}: 过题={nowcoder_record['solved']}, 排名={nowcoder_record['rank']}, 得分={nowcoder_record['score']}")
    
    # 重新计算个人总成绩
    for person in score_data["personal"]:
        team_id = person["team_id"]
        match = re.search(r"team(\d+)", team_id)
        num = int(match.group(1)) if match else 0
        best_n = BEST_N_OLD if num <= TEAM_THRESHOLD else BEST_N_NEW
        
        scores = [c["score"] for c in person["contests"]]
        pairs = [(scores[i], i) for i in range(len(scores))]
        pairs.sort(key=lambda x: x[0], reverse=True)
        top = pairs[:best_n]
        best_indices = set(p[1] for p in top)
        total = sum(p[0] for p in top) / best_n
        
        person["total_score"] = round(total, 2)
        person["best_indices"] = sorted(list(best_indices))
    
    # 重新排序个人
    score_data["personal"].sort(key=lambda x: x["total_score"], reverse=True)
    for i, p in enumerate(score_data["personal"]):
        p["rank"] = i + 1
    
    # 更新队伍总成绩
    name_to_personal = {p["name"]: p for p in score_data["personal"]}
    
    for team in score_data["teams"]:
        member_scores = []
        for member in team["members"]:
            if member["name"] in name_to_personal:
                member["total_score"] = name_to_personal[member["name"]]["total_score"]
                member_scores.append(name_to_personal[member["name"]]["total_score"])
            else:
                member_scores.append(0.0)
        
        team_total = sum(member_scores) / len(member_scores) if member_scores else 0.0
        team["team_total"] = round(team_total, 2)
    
    score_data["teams"].sort(key=lambda x: x["team_total"], reverse=True)
    for i, t in enumerate(score_data["teams"]):
        t["rank"] = i + 1
    
    # 保存更新后的数据
    with open(SCORE_DATA_FILE, "w", encoding="utf-8") as f:
        json.dump(score_data, f, ensure_ascii=False, indent=2)
    
    # 同时更新 public 目录下的文件
    public_file = os.path.join("bupt-ranking", "public", "score_data.json")
    with open(public_file, "w", encoding="utf-8") as f:
        json.dump(score_data, f, ensure_ascii=False, indent=2)
    
    # 同时更新 src/data 目录下的文件
    src_file = os.path.join("bupt-ranking", "src", "data", "score_data.json")
    with open(src_file, "w", encoding="utf-8") as f:
        json.dump(score_data, f, ensure_ascii=False, indent=2)
    
    print("\n成功更新 score_data.json")

if __name__ == "__main__":
    main()