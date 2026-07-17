import json
import re
import os
import time
import pickle
from collections import defaultdict
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from webdriver_manager.chrome import ChromeDriverManager
import openpyxl

CONTEST_ID = "133876"
SEARCH_KEYWORD = "北京邮电大学"
BASELINE = 9  # 牛客第一场 baseline 为 9 题

DATA_DIR = "data"
SCORE_DATA_FILE = os.path.join(DATA_DIR, "score_data.json")
TEAM_FILE = os.path.join(DATA_DIR, "team.xlsx")
OUTPUT_EXCEL = os.path.join(DATA_DIR, "nowcoder_contest1.xlsx")
COOKIE_FILE = "nowcoder_cookies.pkl"

NUM_CONTESTS = 10
BEST_N_OLD = 7
BEST_N_NEW = 5
TEAM_THRESHOLD = 1791


def create_driver():
    options = Options()
    options.add_argument("--disable-blink-features=AutomationControlled")
    options.add_experimental_option("excludeSwitches", ["enable-automation"])
    options.add_experimental_option("useAutomationExtension", False)
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")

    service = Service(ChromeDriverManager().install())
    driver = webdriver.Chrome(service=service, options=options)
    driver.execute_script(
        "Object.defineProperty(navigator, 'webdriver', {get: () => undefined})"
    )
    return driver


def save_cookies(driver, filepath):
    with open(filepath, "wb") as f:
        pickle.dump(driver.get_cookies(), f)
    print(f"Cookies 已保存到 {filepath}")


def load_cookies(driver, filepath):
    if not os.path.exists(filepath):
        return False
    with open(filepath, "rb") as f:
        cookies = pickle.load(f)
    driver.get("https://ac.nowcoder.com/")
    time.sleep(1)
    for cookie in cookies:
        try:
            driver.add_cookie(cookie)
        except Exception:
            pass
    print(f"Cookies 已从 {filepath} 加载")
    return True


def is_logged_in(driver):
    try:
        url = f"https://ac.nowcoder.com/acm/contest/{CONTEST_ID}#rank"
        driver.get(url)
        time.sleep(2)
        current_url = driver.current_url
        if "login" in current_url:
            return False
        return True
    except Exception:
        return False


def wait_for_login(driver):
    url = f"https://ac.nowcoder.com/acm/contest/{CONTEST_ID}#rank"
    driver.get(url)
    print("\n" + "=" * 60)
    print("检测到需要登录，请在浏览器中手动登录牛客账号")
    print("登录成功后脚本会自动继续...")
    print("=" * 60 + "\n")

    while True:
        time.sleep(2)
        current_url = driver.current_url
        if "login" not in current_url:
            print("登录成功！")
            save_cookies(driver, COOKIE_FILE)
            break


def parse_rank_page(driver):
    """解析排名页面，以个人为单位"""
    results = []

    # 等待表格加载
    try:
        WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, "table tbody tr"))
        )
    except Exception:
        print("等待表格超时")
        return results

    time.sleep(2)

    tables = driver.find_elements(By.CSS_SELECTOR, "table")
    target_table = None
    for table in tables:
        rows = table.find_elements(By.CSS_SELECTOR, "tr")
        if len(rows) > 3:
            target_table = table
            break

    if target_table is None:
        print("未找到排名表格")
        return results

    rows = target_table.find_elements(By.CSS_SELECTOR, "tr")

    # 解析表头
    header_row = rows[0]
    header_cells = header_row.find_elements(By.CSS_SELECTOR, "th, td")
    header_texts = [c.text.strip() for c in header_cells]

    print(f"表头: {header_texts}")

    # 查找关键列索引
    rank_idx = None
    name_idx = None
    school_idx = None
    solved_idx = None
    penalty_idx = None

    for i, h in enumerate(header_texts):
        if h == "名次":
            rank_idx = i
        elif h == "参赛者":
            name_idx = i
        elif h == "学校":
            school_idx = i
        elif h == "通过":
            solved_idx = i
        elif h == "罚时":
            penalty_idx = i

    print(f"列索引: 名次={rank_idx}, 参赛者={name_idx}, 学校={school_idx}, 通过={solved_idx}, 罚时={penalty_idx}")

    if name_idx is None or solved_idx is None:
        print("无法定位必要列")
        return results

    # 解析数据行
    for row in rows[1:]:
        cells = row.find_elements(By.CSS_SELECTOR, "td, th")
        if len(cells) <= max(name_idx, solved_idx):
            continue

        # 获取参赛者名称（去除图标等干扰字符）
        name_cell_text = cells[name_idx].text.strip()
        # 清理名称：移除 👤 等图标符号
        name = re.sub(r'[👤\s]+', '', name_cell_text).strip()

        # 获取学校
        school = ""
        if school_idx is not None and len(cells) > school_idx:
            school = cells[school_idx].text.strip()

        # 只保留北京邮电大学的记录
        if "北京邮电大学" not in school and SEARCH_KEYWORD not in school:
            continue

        # 获取排名
        rank_val = ""
        if rank_idx is not None and len(cells) > rank_idx:
            rank_val = cells[rank_idx].text.strip()

        # 获取过题数
        solved_val = ""
        if solved_idx is not None and len(cells) > solved_idx:
            solved_val = cells[solved_idx].text.strip()

        # 获取罚时
        penalty_val = ""
        if penalty_idx is not None and len(cells) > penalty_idx:
            penalty_val = cells[penalty_idx].text.strip()

        if name:
            results.append({
                "name": name,
                "school": school,
                "rank": rank_val,
                "solved": solved_val,
                "penalty": penalty_val,
            })

    return results


def calc_summer_score(solved, rank, baseline):
    """计算暑期训练单场得分
    公式: 得分 = 过题数 / baseline × (601 − 排名) / 600 × 100
    """
    if baseline == 0:
        return 0.0
    try:
        solved_int = int(solved) if solved else 0
        rank_int = int(rank) if rank else 0
    except ValueError:
        return 0.0
    
    score = (solved_int / baseline) * (601 - rank_int) / 600 * 100
    if score < 0 or score > 100:
        return 0.0
    return round(score, 2)


def load_existing_data():
    """加载现有的 score_data.json"""
    if os.path.exists(SCORE_DATA_FILE):
        with open(SCORE_DATA_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    return None


def load_teams():
    """加载队伍信息"""
    wb = openpyxl.load_workbook(TEAM_FILE)
    ws = wb.active
    teams = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        seq, name_cn, name_en, leader, m1, m2 = row
        if name_cn is None:
            continue
        members = [m for m in (leader, m1, m2) if m is not None]
        teams.append({
            "seq": seq,
            "name_cn": name_cn,
            "name_en": name_en,
            "members": members,
        })
    wb.close()
    return teams


def update_score_data(new_contest_data):
    """更新 score_data.json，将牛客第一场成绩加入暑期训练"""
    existing = load_existing_data()
    if not existing:
        print("未找到现有的 score_data.json")
        return

    # 构建现有个人数据的映射（按姓名）
    name_to_personal = {}
    for p in existing["personal"]:
        name_to_personal[p["name"]] = p

    # 更新每个人的第一场成绩
    updated_count = 0
    for item in new_contest_data:
        name = item["name"]
        solved = item["solved"]
        rank = item["rank"]

        # 尝试精确匹配姓名
        person = None
        if name in name_to_personal:
            person = name_to_personal[name]
        else:
            # 尝试模糊匹配（处理名称差异）
            for p_name, p in name_to_personal.items():
                if name in p_name or p_name in name or name.replace(" ", "") == p_name.replace(" ", ""):
                    person = p
                    break

        if person:
            score = calc_summer_score(solved, rank, BASELINE)

            # 更新第一场成绩 (索引 0)
            person["contests"][0] = {
                "solved": int(solved) if solved else 0,
                "rank": int(rank) if rank else 0,
                "score": score,
            }

            # 重新计算个人总成绩
            scores = [c["score"] for c in person["contests"]]
            team_id = person["team_id"]
            match = re.search(r"team(\d+)", team_id)
            num = int(match.group(1)) if match else 0
            best_n = BEST_N_OLD if num <= TEAM_THRESHOLD else BEST_N_NEW

            pairs = [(scores[i], i) for i in range(len(scores))]
            pairs.sort(key=lambda x: x[0], reverse=True)
            top = pairs[:best_n]
            best_indices = set(p[1] for p in top)
            total = sum(p[0] for p in top) / best_n

            person["total_score"] = round(total, 2)
            person["best_indices"] = sorted(list(best_indices))

            updated_count += 1
            print(f"更新 {name}: 过题={solved}, 排名={rank}, 得分={score}")
        else:
            print(f"未匹配到记录: {name}")

    # 重新排序个人
    existing["personal"].sort(key=lambda x: x["total_score"], reverse=True)
    for i, p in enumerate(existing["personal"]):
        p["rank"] = i + 1

    # 更新队伍总成绩
    name_to_personal_updated = {p["name"]: p for p in existing["personal"]}

    for team in existing["teams"]:
        member_scores = []
        for member in team["members"]:
            if member["name"] in name_to_personal_updated:
                member_scores.append(name_to_personal_updated[member["name"]]["total_score"])
                member["total_score"] = name_to_personal_updated[member["name"]]["total_score"]
            else:
                member_scores.append(0.0)

        team_total = sum(member_scores) / len(member_scores) if member_scores else 0.0
        team["team_total"] = round(team_total, 2)

    existing["teams"].sort(key=lambda x: x["team_total"], reverse=True)
    for i, t in enumerate(existing["teams"]):
        t["rank"] = i + 1

    # 保存更新后的数据
    with open(SCORE_DATA_FILE, "w", encoding="utf-8") as f:
        json.dump(existing, f, ensure_ascii=False, indent=2)

    # 同时更新 public 目录下的文件
    public_file = os.path.join("bupt-ranking", "public", "score_data.json")
    with open(public_file, "w", encoding="utf-8") as f:
        json.dump(existing, f, ensure_ascii=False, indent=2)

    # 同时更新 src/data 目录下的文件
    src_file = os.path.join("bupt-ranking", "src", "data", "score_data.json")
    with open(src_file, "w", encoding="utf-8") as f:
        json.dump(existing, f, ensure_ascii=False, indent=2)

    print(f"\n成功更新 {updated_count} 人")
    print(f"已更新文件: {SCORE_DATA_FILE}")


def save_to_excel(results):
    """保存成绩到 Excel 表格"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "牛客第一场成绩"

    # 设置表头
    headers = ["排名", "姓名", "学校", "过题数", "罚时", "Baseline", "得分"]
    ws.append(headers)

    # 设置表头样式
    from openpyxl.styles import Font, Alignment, PatternFill
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # 写入数据
    for item in results:
        score = calc_summer_score(item["solved"], item["rank"], BASELINE)
        ws.append([
            item["rank"],
            item["name"],
            item["school"],
            item["solved"],
            item["penalty"],
            BASELINE,
            score
        ])

    # 设置列宽
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 12

    # 保存
    wb.save(OUTPUT_EXCEL)
    print(f"\n成绩已保存到: {OUTPUT_EXCEL}")


def main():
    print(f"开始爬取牛客竞赛 {CONTEST_ID} 的成绩...")
    print(f"搜索关键词: {SEARCH_KEYWORD}")
    print(f"Baseline: {BASELINE} 题")
    print()

    # 创建驱动
    driver = create_driver()

    try:
        # 尝试加载 cookies
        if not load_cookies(driver, COOKIE_FILE):
            print("未找到缓存的 Cookies")

        # 检查是否已登录
        if not is_logged_in(driver):
            wait_for_login(driver)

        # 访问排名页面
        url = f"https://ac.nowcoder.com/acm/contest/{CONTEST_ID}#rank/%22searchUserName%22%3A%22{SEARCH_KEYWORD}%22"
        print(f"\n访问: {url}")
        driver.get(url)
        time.sleep(3)

        # 解析数据
        results = parse_rank_page(driver)

        if not results:
            print("未找到数据")
            return

        print(f"\n找到 {len(results)} 条个人记录:")
        print("-" * 60)
        for item in results:
            score = calc_summer_score(item["solved"], item["rank"], BASELINE)
            print(f"{item['name']}: 排名={item['rank']}, 过题={item['solved']}, 得分={score}")
        print("-" * 60)

        # 保存到 Excel
        save_to_excel(results)
        print()

        # 确认更新
        confirm = input("是否更新 score_data.json? (y/n): ")
        if confirm.lower() != "y":
            print("已取消更新")
            return

        # 更新数据
        update_score_data(results)
        print("\n更新完成!")

    finally:
        driver.quit()


if __name__ == "__main__":
    main()