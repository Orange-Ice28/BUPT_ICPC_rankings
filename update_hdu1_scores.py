import openpyxl
import json
import sys

sys.stdout.reconfigure(encoding='utf-8')

# 读取Excel数据
wb = openpyxl.load_workbook('data/hdu_contest.xlsx')
ws = wb.active

print(f'Sheet name: {ws.title}')
print(f'Max row: {ws.max_row}, Max col: {ws.max_column}')

# 读取表头
headers = [cell.value for cell in ws[1]]
print(f'Headers: {headers}')

# 读取数据行
hdu1_data = []
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
    if row[0] is not None:  # 姓名不为空
        hdu1_data.append({
            'name': row[0],
            'school': row[1],
            'rank': row[2],
            'solved': row[3],
            'baseline': row[4],
        })

print(f'\n读取到 {len(hdu1_data)} 个队伍的数据')

# 杭电得分公式：得分 = 过题数 / baseline × (501 − 排名) / 500 × 100
def calculate_score(solved, rank, baseline):
    if solved == 0 or solved is None or rank is None or rank == 0:
        return 0.0
    score = (solved / baseline) * ((501 - rank) / 500) * 100
    return round(max(0, score), 2)

# 计算每个队伍的得分
print('\n=== 计算得分 ===')
for item in hdu1_data:
    score = calculate_score(item['solved'], item['rank'], item['baseline'])
    item['score'] = score
    print(f"{item['name']}: 排名={item['rank']}, 过题={item['solved']}, baseline={item['baseline']}, 得分={score}")

# 读取现有JSON数据
with open('bupt-ranking/src/data/summer_score_data.json', 'r', encoding='utf-8') as f:
    data = json.load(f)

# 更新 baseline（杭电1是索引10）
data['baselines'][10] = 7
print(f'\n更新 baselines[10] = 7')

# 队名映射（处理可能的名称差异）
name_mapping = {
    '三只企鹅！': '三只企鹅',
    'ClosedClaw1': 'ClosedClaw',
}

# 更新每个队伍的杭电1数据（索引10）
print('\n=== 更新队伍数据 ===')
for item in hdu1_data:
    team_name = name_mapping.get(item['name'], item['name'])
    
    for team in data['teams']:
        if team['name_cn'] == team_name:
            team['contests'][10] = {
                'solved': item['solved'],
                'rank': item['rank'],
                'score': item['score'],
                'isBest': False
            }
            print(f"Updated: {team['name_cn']} -> solved={item['solved']}, rank={item['rank']}, score={item['score']}")
            break
    else:
        print(f"Warning: 未找到队伍 {team_name}")

# 重新计算每个队伍的总成绩（动态计算有效场次数）
import math

print('\n=== 重新计算总成绩 ===')
for team in data['teams']:
    # 计算有数据的场次数（得分 > 0 或 过题数 > 0）
    contests_with_data = sum(1 for c in team['contests'] if c['solved'] > 0 or c['score'] > 0)
    
    # 初始化所有 contests 的 isBest 为 False
    for contest in team['contests']:
        contest['isBest'] = False
    
    if contests_with_data == 0:
        team['team_total'] = 0.0
    elif contests_with_data == 1:
        # 只有1场时直接计入
        for i, c in enumerate(team['contests']):
            if c['solved'] > 0 or c['score'] > 0:
                team['team_total'] = round(c['score'], 2)
                c['isBest'] = True
                break
    else:
        # 有效场次数 = 有数据的场次 * 80% 向上取整
        best_n = math.ceil(contests_with_data * 0.8)
        # 获取有数据的场次索引和分数
        contest_pairs = [(i, c['score']) for i, c in enumerate(team['contests']) if c['solved'] > 0 or c['score'] > 0]
        # 按分数降序排序
        contest_pairs.sort(key=lambda x: x[1], reverse=True)
        # 取最好N场
        best_pairs = contest_pairs[:best_n]
        best_indices = set(p[0] for p in best_pairs)
        # 标记 best contests
        for i in best_indices:
            team['contests'][i]['isBest'] = True
        
        best_scores = [p[1] for p in best_pairs]
        team['team_total'] = round(sum(best_scores) / best_n, 2)
    
    print(f"{team['name_cn']}: contests_with_data={contests_with_data}, best_n={math.ceil(contests_with_data * 0.8) if contests_with_data > 1 else contests_with_data}, team_total={team['team_total']}")

# 重新排名
data['teams'].sort(key=lambda x: x['team_total'], reverse=True)
for i, team in enumerate(data['teams']):
    team['rank'] = i + 1

# 保存更新后的数据
with open('bupt-ranking/src/data/summer_score_data.json', 'w', encoding='utf-8') as f:
    json.dump(data, f, ensure_ascii=False, indent=2)

print('\n✓ 数据已更新并保存到 bupt-ranking/src/data/summer_score_data.json')